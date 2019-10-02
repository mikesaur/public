from openpyxl import load_workbook
from sys import argv

testbed_name = argv[1]
excel_file = argv[2]

wb = load_workbook(excel_file)
ws = wb[wb.sheetnames[0]]

customer_devices = open(testbed_name + '_devices.yaml', 'w')  # Create file

username = argv[3]
password = argv[4]

customer_devices.write('testbed:\n')
customer_devices.write('\n')
customer_devices.write('   name: ' + testbed_name + '\n')
customer_devices.write('\n')
customer_devices.write('\n')
customer_devices.write('devices:\n')

for row in range(2, 50):
    router_name = ws.cell(row=row, column=1)
    router_ip = ws.cell(row=row, column=2)
    router_type = ws.cell(row=row, column=3)
    if router_name.value is not None:
        customer_devices.write('  ' + router_name.value + ':\n')
        customer_devices.write('    alias: ' + router_name.value + '\n')
        customer_devices.write('    os: ios' + str(router_type.value) + '\n')
        customer_devices.write('    type: router\n')
        customer_devices.write('    tacacs:\n')
        customer_devices.write('        username: ' + username + ' \n')
        customer_devices.write('    passwords:\n')
        customer_devices.write('        tacacs: ' + password + ' \n')
        customer_devices.write('    connections:\n')
        customer_devices.write('      defaults:\n')
        customer_devices.write('        class: unicon.Unicon\n')
        customer_devices.write('      console:\n')
        customer_devices.write('        ip: ' + router_ip.value + '\n')
        customer_devices.write('        protocol: ssh\n')
        customer_devices.write('        port: 22\n')
customer_devices.close()

print('Yaml file built!!')
